from __future__ import annotations

import csv
import io
import logging
import os
import tempfile
from datetime import datetime, timezone

from aiogram import Bot, Router
from aiogram.filters import Command
from aiogram.types import BufferedInputFile, Message
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ai_analyzer import AIAnalyzer
from db import Database
from keyboards import (
    PDF_SLICE_SIZE,
    admin_menu_keyboard,
    panel_keyboard,
    pdf_size_keyboard,
    status_keyboard,
)
from pdf_export import build_bugs_pdf
from persian import fa_digits, media_label, to_shamsi

logger = logging.getLogger(__name__)

router = Router(name="admin")

# Telegram photo captions are capped at 1024 chars; keep the inline raw-text
# excerpt well under that so a photo card never gets rejected.
_CARD_TEXT_LIMIT = 400


def _admin_filter_factory(super_admin_ids: tuple[int, ...], db: Database):
    async def is_admin(message: Message) -> bool:
        user = message.from_user
        if user is None:
            return False
        if user.id in super_admin_ids:
            return True
        return await db.is_admin(user.id)

    return is_admin


def setup(super_admin_ids: tuple[int, ...], db: Database) -> Router:
    router.message.filter(_admin_filter_factory(super_admin_ids, db))
    return router


def format_bug(bug: dict) -> str:
    """Persian, right-to-left bug card. Telegram renders Persian text RTL
    automatically; we lead each line with its label so it reads cleanly."""
    lines = [
        f"🐞 گزارش BUG-{fa_digits(bug['id'])}",
        f"📅 تاریخ ثبت: {to_shamsi(bug.get('created_at'))}",
        f"👤 گزارش‌دهنده: {bug.get('reporter_name') or '—'}",
        f"📝 عنوان: {bug.get('ai_title') or '—'}",
        f"🏷 دسته: {bug.get('category') or '—'}",
        f"🔺 شدت: {bug.get('severity') or '—'}",
        f"📌 وضعیت: {bug.get('status') or '—'}",
    ]
    if bug.get("telegram_file_id"):
        lines.append(f"📎 فایل پیوست: {media_label(bug.get('media_type'))}")
    if bug.get("ai_status") == "PENDING":
        lines.append("⏳ هوش مصنوعی: هنوز بررسی نشده")

    raw = (bug.get("raw_text") or "").strip()
    if raw:
        excerpt = raw if len(raw) <= _CARD_TEXT_LIMIT else raw[:_CARD_TEXT_LIMIT] + "…"
        lines.append(f"\n📄 متن گزارش:\n{excerpt}")
    return "\n".join(lines)


async def send_bug_card(
    message: Message, bug: dict, is_super: bool = False
) -> None:
    """Send one bug as a card. For photo reports the screenshot is shown inline
    (card text as the caption); other media keep the 📎 view button."""
    kb = status_keyboard(
        bug["id"],
        has_media=bool(bug.get("telegram_file_id")),
        is_super=is_super,
        media_type=bug.get("media_type"),
    )
    if bug.get("media_type") == "photo" and bug.get("telegram_file_id"):
        try:
            await message.answer_photo(
                bug["telegram_file_id"],
                caption=format_bug(bug),
                reply_markup=kb,
            )
            return
        except Exception as exc:
            logger.warning("Inline photo for BUG-%s failed: %s", bug["id"], exc)
    await message.answer(format_bug(bug), reply_markup=kb)


def _rows_for_export(bugs: list[dict]) -> list[list]:
    rows: list[list] = []
    for b in bugs:
        has_file = bool(b.get("telegram_file_id"))
        rows.append(
            [
                fa_digits(b["id"]),
                to_shamsi(b.get("created_at")),
                b.get("reporter_name") or "—",
                b.get("ai_title") or "—",
                b.get("category") or "—",
                b.get("severity") or "—",
                b.get("priority") or "—",
                b.get("suggested_owner") or "—",
                b.get("status") or "—",
                media_label(b.get("media_type")) if has_file else "—",
                b.get("ai_summary") or "—",
                b.get("raw_text") or "—",
            ]
        )
    return rows


EXPORT_HEADERS = [
    "شماره",
    "تاریخ ثبت",
    "گزارش‌دهنده",
    "عنوان",
    "دسته",
    "شدت",
    "اولویت",
    "مسئول",
    "وضعیت",
    "فایل پیوست",
    "خلاصه",
    "متن کامل گزارش",
]


async def send_csv(message: Message, db: Database) -> None:
    bugs = await db.list_all()
    if not bugs:
        await message.answer("⚠️ هیچ گزارشی برای خروجی گرفتن نیست.")
        return

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(EXPORT_HEADERS)
    for row in _rows_for_export(bugs):
        writer.writerow(row)
    data = buf.getvalue().encode("utf-8-sig")
    name = f"bugs-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.csv"
    await message.answer_document(BufferedInputFile(data, filename=name))


# Columns (0-based) whose content is long free text — they get a fixed wide
# width with wrapping instead of stretching to the longest line.
_WRAP_COLUMNS = {3, 10, 11}  # عنوان، خلاصه، متن کامل گزارش
_WRAP_WIDTH = 50


async def send_xlsx(message: Message, db: Database) -> None:
    bugs = await db.list_all()
    if not bugs:
        await message.answer("⚠️ هیچ گزارشی برای خروجی گرفتن نیست.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش‌ها"
    ws.sheet_view.rightToLeft = True  # RTL sheet

    ws.append(EXPORT_HEADERS)
    for row in _rows_for_export(bugs):
        ws.append(row)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="right", vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = cell_align

    for idx, col_cells in enumerate(ws.columns):
        letter = col_cells[0].column_letter
        if idx in _WRAP_COLUMNS:
            ws.column_dimensions[letter].width = _WRAP_WIDTH
        else:
            values = [c.value for c in col_cells if c.value is not None]
            max_len = max((len(str(v)) for v in values), default=10)
            ws.column_dimensions[letter].width = min(max_len + 2, 30)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    name = f"bugs-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"
    await message.answer_document(BufferedInputFile(buf.getvalue(), filename=name))


async def _build_backup(db: Database) -> tuple[bytes, str]:
    """Snapshot the live DB into a temp file and return its bytes + filename."""
    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    fname = f"bugs-backup-{ts}.db"
    tmp_path = os.path.join(tempfile.gettempdir(), fname)
    await db.backup_to(tmp_path)
    try:
        with open(tmp_path, "rb") as fh:
            data = fh.read()
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
    return data, fname


async def send_backup(bot: Bot, db: Database, chat_id: int) -> None:
    data, fname = await _build_backup(db)
    await bot.send_document(
        chat_id,
        BufferedInputFile(data, filename=fname),
        caption=f"🗄 بکاپ دیتابیس\n{fname}",
    )


async def broadcast_backup(bot: Bot, db: Database, chat_ids: tuple[int, ...]) -> None:
    """Build one snapshot and send it to every super admin (used by the
    daily auto-backup loop)."""
    if not chat_ids:
        return
    data, fname = await _build_backup(db)
    for cid in chat_ids:
        try:
            await bot.send_document(
                cid,
                BufferedInputFile(data, filename=fname),
                caption=f"🗄 بکاپ خودکار روزانه\n{fname}",
            )
        except Exception as exc:
            logger.warning("Daily backup send to %s failed: %s", cid, exc)


async def reanalyze_pending(message: Message, db: Database, analyzer: AIAnalyzer) -> None:
    pending = await db.list_pending(200)
    if not pending:
        await message.answer("✅ همه‌ی گزارش‌ها تحلیل شدن. چیزی برای بررسی نیست.")
        return

    status_msg = await message.answer(
        f"♻️ در حال تحلیل {len(pending)} گزارش بررسی‌نشده..."
    )
    ok_count = 0
    fail_count = 0
    for bug in pending:
        analysis, ai_ok = await analyzer.analyze(
            bug.get("raw_text") or "", bug.get("media_type")
        )
        if ai_ok:
            await db.update_analysis(bug["id"], analysis, "ANALYZED")
            ok_count += 1
        else:
            fail_count += 1

    result = f"✅ {ok_count} گزارش با موفقیت تحلیل شد."
    if fail_count:
        result += (
            f"\n⚠️ {fail_count} گزارش هنوز ناموفق بود "
            "(احتمالاً اعتبار OpenAI هنوز مشکل داره)."
        )
    try:
        await status_msg.edit_text(result)
    except Exception:
        await message.answer(result)


async def send_panel(message: Message, db: Database) -> None:
    s = await db.stats()
    text = (
        "📊 داشبورد گزارش‌ها\n\n"
        f"📥 کل گزارش‌ها: {fa_digits(s['total'])}\n"
        f"🟢 باز: {fa_digits(s['open'])}\n"
        f"⏳ منتظر تحلیل: {fa_digits(s['pending'])}\n\n"
        f"🔴 Critical: {fa_digits(s['Critical'])}\n"
        f"🟠 High: {fa_digits(s['High'])}\n"
        f"🟡 Medium: {fa_digits(s['Medium'])}\n"
        f"🟢 Low: {fa_digits(s['Low'])}"
    )
    await message.answer(text, reply_markup=panel_keyboard())


async def admins_text(db: Database, super_admin_ids: tuple[int, ...]) -> str:
    lines = ["👑 سوپرادمین‌ها:"]
    for sid in super_admin_ids:
        lines.append(f"• `{sid}`")

    admins = await db.list_admins()
    lines.append("\n👤 ادمین‌ها:")
    if admins:
        for a in admins:
            nm = f" ({a['name']})" if a.get("name") else ""
            lines.append(f"• `{a['user_id']}`{nm}")
    else:
        lines.append("(هنوز ادمینی اضافه نشده — با /add_admin اضافه کن)")
    return "\n".join(lines)


@router.message(Command("panel"))
async def cmd_panel(message: Message, db: Database) -> None:
    await send_panel(message, db)


@router.message(Command("reanalyze"))
async def cmd_reanalyze(message: Message, db: Database, analyzer: AIAnalyzer) -> None:
    await reanalyze_pending(message, db, analyzer)


@router.message(Command("bugs"))
async def cmd_bugs(
    message: Message, db: Database, super_admin_ids: tuple[int, ...] = ()
) -> None:
    bugs = await db.list_latest(20)
    if not bugs:
        await message.answer("هنوز هیچ باگی ثبت نشده.")
        return

    is_super = _is_super(message, super_admin_ids)
    await message.answer(f"📋 {fa_digits(len(bugs))} گزارش آخر:")
    for bug in bugs:
        await send_bug_card(message, bug, is_super)


@router.message(Command("export_csv"))
async def cmd_export_csv(message: Message, db: Database) -> None:
    await send_csv(message, db)


@router.message(Command("export_excel"))
async def cmd_export_excel(message: Message, db: Database) -> None:
    await send_xlsx(message, db)


async def send_pdf_for_bugs(message: Message, bugs: list[dict]) -> None:
    """Build a PDF for the supplied bug rows and deliver it to the chat.
    Caller is responsible for any progress message that goes with it."""
    if not bugs:
        await message.answer("⚠️ هیچ گزارشی برای خروجی گرفتن نیست.")
        return
    if message.bot is None:
        await message.answer("❌ خطا در دسترسی به ربات.")
        return
    try:
        data = await build_bugs_pdf(message.bot, bugs)
    except Exception as exc:
        logger.exception("PDF export failed: %s", exc)
        await message.answer("❌ ساخت PDF ناموفق بود.")
        return
    name = f"bugs-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.pdf"
    await message.answer_document(BufferedInputFile(data, filename=name))


async def send_pdf(message: Message, db: Database) -> None:
    """For small exports, build the PDF directly. For large ones, ask the user
    how many of the latest reports to include (or 'all')."""
    total = (await db.stats())["total"]
    if total == 0:
        await message.answer("⚠️ هیچ گزارشی برای خروجی گرفتن نیست.")
        return
    if total <= PDF_SLICE_SIZE:
        await message.answer(
            "📕 در حال ساخت PDF... (چند ثانیه برای دانلود عکس‌ها)"
        )
        await send_pdf_for_bugs(message, await db.list_all())
        return
    await message.answer(
        "📕 خروجی PDF\n\n"
        f"📥 کل گزارش‌ها: {fa_digits(total)}\n"
        "چند گزارش توی PDF بیاد؟",
        reply_markup=pdf_size_keyboard(total),
    )


@router.message(Command("export_pdf"))
async def cmd_export_pdf(message: Message, db: Database) -> None:
    await send_pdf(message, db)


# ----- Super-admin-only commands ------------------------------------------------


def _is_super(message: Message, super_admin_ids: tuple[int, ...]) -> bool:
    return bool(message.from_user and message.from_user.id in super_admin_ids)


@router.message(Command("add_admin"))
async def cmd_add_admin(
    message: Message, db: Database, super_admin_ids: tuple[int, ...] = ()
) -> None:
    if not _is_super(message, super_admin_ids):
        await message.answer("⛔️ فقط سوپرادمین می‌تونه ادمین اضافه کنه.")
        return

    parts = (message.text or "").split()
    if len(parts) < 2:
        await message.answer(
            "استفاده:\n`/add_admin <آیدی عددی> [نام دلخواه]`\n\n"
            "ادمین جدید آیدی عددیش رو با دستور /myid می‌تونه ببینه."
        )
        return
    try:
        new_id = int(parts[1])
    except ValueError:
        await message.answer("❌ آیدی باید یک عدد باشه.")
        return

    if new_id in super_admin_ids:
        await message.answer("این کاربر سوپرادمینه و از قبل دسترسی کامل داره.")
        return

    name = " ".join(parts[2:]).strip()
    await db.add_admin(new_id, name=name, added_by=message.from_user.id)
    suffix = f" ({name})" if name else ""
    await message.answer(
        f"✅ ادمین جدید اضافه شد: `{new_id}`{suffix}\n"
        "حالا می‌تونه /panel و /bugs و خروجی‌ها رو ببینه."
    )


@router.message(Command("remove_admin"))
async def cmd_remove_admin(
    message: Message, db: Database, super_admin_ids: tuple[int, ...] = ()
) -> None:
    if not _is_super(message, super_admin_ids):
        await message.answer("⛔️ فقط سوپرادمین می‌تونه ادمین حذف کنه.")
        return

    parts = (message.text or "").split()
    if len(parts) < 2:
        await message.answer("استفاده:\n`/remove_admin <آیدی عددی>`")
        return
    try:
        rid = int(parts[1])
    except ValueError:
        await message.answer("❌ آیدی باید یک عدد باشه.")
        return

    ok = await db.remove_admin(rid)
    await message.answer(
        f"✅ ادمین `{rid}` حذف شد." if ok else "این آیدی توی لیست ادمین‌ها نبود."
    )


@router.message(Command("admin"))
async def cmd_admin(
    message: Message, db: Database, super_admin_ids: tuple[int, ...] = ()
) -> None:
    if not _is_super(message, super_admin_ids):
        await message.answer(
            "⛔️ منوی /admin فقط برای سوپرادمینه.\n"
            "برای دیدن گزارش‌ها از دستور /panel استفاده کن."
        )
        return
    await message.answer(
        "🛠 پنل سوپرادمین\nیک گزینه رو انتخاب کن:",
        reply_markup=admin_menu_keyboard(),
    )


@router.message(Command("admins"))
async def cmd_admins(
    message: Message, db: Database, super_admin_ids: tuple[int, ...] = ()
) -> None:
    if not _is_super(message, super_admin_ids):
        await message.answer("⛔️ فقط سوپرادمین به این دستور دسترسی داره.")
        return
    await message.answer(await admins_text(db, super_admin_ids))


@router.message(Command("backup"))
async def cmd_backup(
    message: Message, db: Database, super_admin_ids: tuple[int, ...] = ()
) -> None:
    if not _is_super(message, super_admin_ids):
        await message.answer("⛔️ فقط سوپرادمین می‌تونه بکاپ بگیره.")
        return
    if message.bot is None:
        return
    try:
        await message.answer("🗄 در حال ساخت بکاپ...")
        await send_backup(message.bot, db, message.chat.id)
    except Exception as exc:
        logger.exception("Manual backup failed: %s", exc)
        await message.answer("❌ ساخت بکاپ ناموفق بود.")


@router.message(Command("delete"))
async def cmd_delete(
    message: Message, db: Database, super_admin_ids: tuple[int, ...] = ()
) -> None:
    if not _is_super(message, super_admin_ids):
        await message.answer("⛔️ فقط سوپرادمین می‌تونه گزارش حذف کنه.")
        return

    parts = (message.text or "").split()
    if len(parts) < 2:
        await message.answer(
            "استفاده:\n`/delete <شماره گزارش>`\n"
            "مثال: `/delete 12`\n\n"
            "یا توی /bugs روی دکمهٔ «🗑 حذف گزارش» بزن."
        )
        return

    raw = parts[1].strip().lstrip("#")
    raw = raw.replace("BUG-", "").replace("bug-", "")
    try:
        bug_id = int(raw)
    except ValueError:
        await message.answer("❌ شماره گزارش باید عدد باشه.")
        return

    ok = await db.delete_bug(bug_id)
    await message.answer(
        f"🗑 BUG-{bug_id} حذف شد." if ok else f"گزارشی با شماره BUG-{bug_id} پیدا نشد."
    )
